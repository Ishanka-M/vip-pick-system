"""
pick_engine.py — Inventory matching, pallet-level allocation, WMS output build
=============================================================================
Rules
  * Item match  : base ID only (P162400-000-140 -> P162400)
  * Plant       : user confirm කරපු plant(s) එකෙන් විතරයි
  * Whole doc   : Invoice / DC එකේ හැම line එකකටම stock තිබ්බොත් විතරයි pick
  * Duplicate   : එකම Invoice / DC number එකකට එක output එකයි
  * Pallet level: කොච්චර pick කරාද, balance කීයද — හැම pallet එකකටම save
"""
from __future__ import annotations

import io
import re
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from doc_parser import ParsedDoc, base_item, clean_item

# Bumped whenever this module's public surface changes; app.py refuses to run
# against a stale copy instead of dying with a redacted TypeError.
API = 8

# --------------------------------------------------------------------------- #
# WMS templates
# --------------------------------------------------------------------------- #
MASTER_COLS = [
    "HOST_ORDER_MASTER_ID", "HOST_GROUP_ID", "RECORD_CREATE_DATE", "PROCESSING_CODE",
    "WH_ID", "CLIENT_CODE", "ORDER_NUMBER", "DISPLAY_ORDER_NUMBER", "STORE_ORDER_NUMBER",
    "ORDER_TYPE", "CUSTOMER_CODE", "CUSTOMER_PO_NUMBER", "DEPARTMENT", "LOAD_ID",
    "LOAD_SEQ", "BOL_NUMBER", "MASTER_BOL_NUMBER", "PRO_NUMBER", "CARRIER", "CARRIER_SCAC",
    "FREIGHT_TERMS", "RUSH", "ORDER_DATE", "ARRIVE_DATE", "DATE_EXPECTED", "PROMISED_DATE",
    "WEIGHT", "CUBIC_VOLUME", "CONTAINERS", "BACKORDER", "PRE_PAID", "COD_AMOUNT",
    "INSURANCE_AMOUNT", "PIP_AMOUNT", "FREIGHT_COST", "SHIP_TO_CODE", "SHIP_TO_NAME",
    "SHIP_TO_ADDR1", "SHIP_TO_ADDR2", "SHIP_TO_ADDR3", "SHIP_TO_CITY", "SHIP_TO_STATE",
    "SHIP_TO_ZIP", "SHIP_TO_COUNTRY_CODE", "SHIP_TO_COUNTRY_NAME", "SHIP_TO_PHONE",
    "BILL_TO_CODE", "BILL_TO_NAME", "BILL_TO_ADDR1", "BILL_TO_ADDR2", "BILL_TO_ADDR3",
    "BILL_TO_CITY", "BILL_TO_STATE", "BILL_TO_ZIP", "BILL_TO_COUNTRY_CODE",
    "BILL_TO_COUNTRY_NAME", "BILL_TO_PHONE", "DELIVERY_NAME", "DELIVERY_ADDR1",
    "DELIVERY_ADDR2", "DELIVERY_ADDR3", "DELIVERY_CITY", "DELIVERY_STATE", "DELIVERY_ZIP",
    "DELIVERY_COUNTRY_CODE", "DELIVERY_COUNTRY_NAME", "DELIVERY_PHONE",
    "BILL_FRGHT_TO_CODE", "BILL_FRGHT_TO_NAME", "BILL_FRGHT_TO_ADDR1",
    "BILL_FRGHT_TO_ADDR2", "BILL_FRGHT_TO_ADDR3", "BILL_FRGHT_TO_CITY",
    "BILL_FRGHT_TO_STATE", "BILL_FRGHT_TO_ZIP", "BILL_FRGHT_TO_COUNTRY_CODE",
    "BILL_FRGHT_TO_COUNTRY_NAME", "BILL_FRGHT_TO_PHONE", "CARTON_LABEL", "VER_FLAG",
    "PARTIAL_ORDER_FLAG", "EARLIEST_SHIP_DATE", "LATEST_SHIP_DATE",
    "EARLIEST_DELIVERY_DATE", "LATEST_DELIVERY_DATE", "TEMP_LINK_ID", "SERVICE_LEVEL",
    "SHIP_VIA", "SHIP_TO_ATTENTION", "SAT_DELIVERY_FLAG", "REGISTERED_MAIL_FLAG",
    "RESTRICTED_MAIL_FLAG", "COD_FLAG", "COD_PAY_TYPE", "COD_OPTION", "INSURANCE_FLAG",
    "BILL_FRGHT_TO_ATTENTION", "SHIP_TO_RESIDENTIAL_FLAG", "CARRIER_MODE",
    "EARLIEST_APPT_TIME", "LATEST_APPT_TIME",
]

DETAIL_COLS = [
    "HOST_ORDER_DETAIL_ID", "HOST_ORDER_MASTER_ID", "HOST_GROUP_ID", "RECORD_CREATE_DATE",
    "PROCESSING_CODE", "WH_ID", "CLIENT_CODE", "ORDER_NUMBER", "DISPLAY_ORDER_NUMBER",
    "LINE_NUMBER", "ITEM_NUMBER", "DISPLAY_ITEM_NUMBER", "ITEM_DESCRIPTION", "CUST_PART",
    "LOT_NUMBER", "QTY", "UNIT_WEIGHT", "UNIT_VOLUME", "EXTENDED_WEIGHT",
    "EXTENDED_VOLUME", "HAZ_MATERIAL", "BOL_CLASS", "BOL_CODE", "ORDER_UOM",
    "HOST_WAVE_ID", "TEMP_LINK_ID", "UNIT_INSURANCE_AMOUNT",
    "GEN_ATTRIBUTE_VALUE1", "GEN_ATTRIBUTE_VALUE2", "GEN_ATTRIBUTE_VALUE3",
    "GEN_ATTRIBUTE_VALUE4", "GEN_ATTRIBUTE_VALUE5", "GEN_ATTRIBUTE_VALUE6",
    "GEN_ATTRIBUTE_VALUE7", "GEN_ATTRIBUTE_VALUE8", "GEN_ATTRIBUTE_VALUE9",
    "GEN_ATTRIBUTE_VALUE10", "GEN_ATTRIBUTE_VALUE11", "HOLD_REASON_ID", "PACKING_INST",
    "REIMA_LINE", "VAS_INST",
]

MASTER_FIXED = {
    "PROCESSING_CODE": "NEW", "ORDER_TYPE": "Sales Orders", "BACKORDER": "N",
    "PARTIAL_ORDER_FLAG": "N", "SAT_DELIVERY_FLAG": "N", "REGISTERED_MAIL_FLAG": "N",
    "RESTRICTED_MAIL_FLAG": "N", "COD_FLAG": "N", "COD_PAY_TYPE": "N", "COD_OPTION": "N",
    "INSURANCE_FLAG": "N", "SHIP_TO_RESIDENTIAL_FLAG": "N",
}
DETAIL_FIXED = {"PROCESSING_CODE": "NEW"}

# GEN_ATTRIBUTE_VALUEn  ->  Inventory column
GEN_MAP = {
    "GEN_ATTRIBUTE_VALUE1": "color",
    "GEN_ATTRIBUTE_VALUE2": "size",
    "GEN_ATTRIBUTE_VALUE3": "style",
    "GEN_ATTRIBUTE_VALUE4": "supplier",
    "GEN_ATTRIBUTE_VALUE5": "plant",
    "GEN_ATTRIBUTE_VALUE6": "client_so",
    "GEN_ATTRIBUTE_VALUE7": "client_so_line",
    "GEN_ATTRIBUTE_VALUE8": "po_cust_dec",
    "GEN_ATTRIBUTE_VALUE9": "customer_ref_number",
    "GEN_ATTRIBUTE_VALUE10": "item_id",
    "GEN_ATTRIBUTE_VALUE11": "invoice_number_1",
}

ALLOC_COLS = [
    "RUN_ID", "PICK_DATE", "DOC_TYPE", "DOC_NUMBER", "DOC_LINE", "DOC_ITEM_CODE",
    "BASE_ID", "ITEM_NUMBER", "DESCRIPTION", "PALLET", "LOCATION_ID", "LOT_NUMBER",
    "PLANT", "UOM", "QTY_BEFORE", "QTY_PICKED", "QTY_BALANCE", "FIFO_DATE",
    "GRN_NUMBER", "STORED_ATTRIBUTE_ID", "ROW_KEY", "SOURCE_FILE",
]

# --------------------------------------------------------------------------- #
# Inventory normalisation
# --------------------------------------------------------------------------- #
INV_ALIASES: dict[str, list[str]] = {
    "wh_id": ["wh id", "warehouse id"],
    "client_code": ["client code"],
    "pallet": ["pallet", "pallet id", "lpn"],
    "location_id": ["location id", "location"],
    "item_number": ["item number", "item no"],
    "display_item_number": ["display item number"],
    "description": ["description", "item description"],
    "lot_number": ["lot number", "lot"],
    "actual_qty": ["actual qty", "available qty", "qty", "quantity"],
    "unavailable_qty": ["unavailable qty", "unavailable"],
    "uom": ["uom", "unit of measure"],
    "status": ["status"],
    "pick_id": ["pick id", "pickid", "pick task id", "pick task"],
    "stored_attribute_id": ["stored attribute id"],
    "fifo_date": ["fifo date"],
    "grn_number": ["grn number", "grn"],
    "cbm": ["cbm"],
    "color": ["color", "colour"],
    "size": ["size"],
    "style": ["style"],
    "supplier": ["supplier"],
    "plant": ["plant"],
    "client_so": ["client so"],
    "client_so_line": ["client so line"],
    "po_cust_dec": ["po cust dec"],
    "customer_ref_number": ["customer ref number"],
    "item_id": ["item id"],
    "invoice_number_1": ["invoice number 1"],
}


def _norm(s: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def normalize_inventory(df: pd.DataFrame) -> pd.DataFrame:
    """Inventory report -> canonical columns + numeric qty + base id."""
    src = {c: _norm(c) for c in df.columns}
    used: set[str] = set()
    out = pd.DataFrame(index=df.index)

    for canon, aliases in INV_ALIASES.items():
        found = None
        for a in aliases:                                   # exact match first
            na = _norm(a)
            for col, nc in src.items():
                if nc == na and col not in used:
                    found = col
                    break
            if found:
                break
        if found is None:                                   # then contains
            for a in aliases:
                na = _norm(a)
                for col, nc in src.items():
                    if na and na in nc and col not in used:
                        found = col
                        break
                if found:
                    break
        if found is not None:
            used.add(found)
            out[canon] = df[found]
        else:
            out[canon] = ""

    out["actual_qty"] = pd.to_numeric(out["actual_qty"], errors="coerce").fillna(0.0)
    out["unavailable_qty"] = pd.to_numeric(out["unavailable_qty"], errors="coerce").fillna(0.0)
    # raw = WMS එකේ තියෙන හරියටම item number (trailing '.' වගේ ඒවත් එහෙමම තියෙන්න ඕන)
    out["item_number_raw"] = out["item_number"].astype(str).str.strip().replace(
        {"nan": "", "None": ""})
    out["item_number"] = out["item_number"].map(clean_item)      # matching key
    # base_id must come off the *raw* code: clean_item() deletes the space in
    # "P601560 710", and base_item() needs that separator to see 710 as a
    # suffix. Deriving it from the cleaned key gave "P601560710" here while the
    # document side gave "P601560" — the same item never matched, and the
    # document was rejected as "Item not in inventory".
    out["base_id"] = out["item_number_raw"].map(base_item)
    out["_fifo"] = pd.to_datetime(out["fifo_date"], errors="coerce", dayfirst=True)
    for c in ("pallet", "location_id", "lot_number", "plant", "uom", "status",
              "stored_attribute_id"):
        out[c] = out[c].astype(str).replace({"nan": "", "None": "", "NaT": ""}).str.strip()

    # Pick Id — 0 කියන්නේ free. 0 නොවන එකක් තියෙනවා නම් ඒ pallet එක දැනටමත්
    # WMS එකේ pick task එකකට allocate වෙලා (Status එක තාම 'Available' වුණත්).
    out["pick_id"] = (out["pick_id"].astype(str).str.strip()
                      .str.replace(r"\.0$", "", regex=True)
                      .replace({"nan": "", "None": "", "NaT": "", "-": ""}))
    out["pick_free"] = out["pick_id"].isin(["", "0"])
    out["row_key"] = (
        out["pallet"] + "|" + out["location_id"] + "|" + out["item_number"] + "|"
        + out["lot_number"] + "|" + out["stored_attribute_id"].astype(str)
    )
    out["free_qty"] = (out["actual_qty"] - out["unavailable_qty"]).clip(lower=0)
    out = out[out["item_number"] != ""].reset_index(drop=True)
    return out


def plant_summary(inv: pd.DataFrame) -> pd.DataFrame:
    free = inv[inv["pick_free"]] if "pick_free" in inv.columns else inv
    g = (free.groupby("plant", dropna=False)
             .agg(Pallets=("pallet", "nunique"), Items=("item_number", "nunique"),
                  Qty=("free_qty", "sum"))
             .reset_index().rename(columns={"plant": "Plant"}))
    g["Qty"] = g["Qty"].astype(int)
    if "pick_free" in inv.columns:
        lock = (inv[~inv["pick_free"]].groupby("plant")["free_qty"].sum()
                .rename("On pick task"))
        g["On pick task"] = (g["Plant"].map(lock).fillna(0).astype(int))
    return g.sort_values("Qty", ascending=False).reset_index(drop=True)


# --------------------------------------------------------------------------- #
# Pallet-level stock basis
# --------------------------------------------------------------------------- #
QTY_TOL = 1e-6

BASIS_COLS = ["PALLET", "LOCATION_ID", "ITEM_NUMBER", "BASE_ID", "LOT_NUMBER", "PLANT",
              "UOM", "PICK_ID", "PICK_STATUS", "INV_ACTUAL_QTY", "LEDGER_BEFORE",
              "LEDGER_PICKED", "LEDGER_BALANCE", "MODE", "AVAILABLE", "ROW_KEY"]


def _soft_key(pallet: Any, item: Any, lot: Any) -> str:
    return (f"{str(pallet or '').strip().upper()}|{clean_item(item)}|"
            f"{str(lot or '').strip().upper()}")


def ledger_state(ledger: pd.DataFrame | None) -> tuple[dict[str, dict], dict[str, dict]]:
    """
    PALLET_LEDGER -> pallet+item එකකට { before, picked, balance }.
    before  = ledger එකේ තියෙන ලොකුම QTY_BEFORE (chain එකේ මුල් baseline එක)
    picked  = ඒ pallet එකෙන් මුළු pick කරපු ප්‍රමාණය
    (exact ROW_KEY dict, soft PALLET|ITEM|LOT dict) විදිහට දෙකක් return වෙනවා.
    """
    empty: dict[str, dict] = {}
    if ledger is None or not len(ledger) or "ROW_KEY" not in ledger.columns:
        return empty, empty

    led = ledger.copy()
    for c in ("QTY_BEFORE", "QTY_PICKED", "QTY_BALANCE"):
        led[c] = pd.to_numeric(led.get(c), errors="coerce").fillna(0.0)
    led["ROW_KEY"] = led["ROW_KEY"].astype(str)
    led["_soft"] = [
        _soft_key(p, i, l) for p, i, l in zip(
            led.get("PALLET", ""), led.get("ITEM_NUMBER", ""), led.get("LOT_NUMBER", ""))
    ]

    def _pack(gcol: str) -> dict[str, dict]:
        g = led.groupby(gcol).agg(before=("QTY_BEFORE", "max"),
                                  picked=("QTY_PICKED", "sum"),
                                  last=("QTY_BALANCE", "min"))
        out: dict[str, dict] = {}
        for k, r in g.iterrows():
            before, picked = float(r["before"]), float(r["picked"])
            out[str(k)] = {"before": before, "picked": picked,
                           "balance": max(0.0, before - picked)}
        return out

    return _pack("ROW_KEY"), _pack("_soft")


def stock_basis(inv: pd.DataFrame, ledger: pd.DataFrame | None = None,
                use_ledger: bool = True) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    හැම pallet+item එකකටම pick කරන්න පුළුවන් ප්‍රමාණය තීරණය කරනවා:

      * Ledger එකේ නෑ                      -> AVAILABLE = Inventory Actual Qty
      * Actual Qty == ledger QTY_BEFORE     -> Inventory refresh වෙලා නෑ
                                               -> AVAILABLE = QTY_BALANCE
      * Actual Qty != ledger QTY_BEFORE     -> WMS update වෙලා, අලුත් baseline
                                               -> AVAILABLE = Actual Qty

    Return: (inv rows + 'avail' column, basis DataFrame)
    """
    exact, soft = ledger_state(ledger) if use_ledger else ({}, {})

    agg = (inv.groupby("row_key", as_index=False)["free_qty"].sum()
             .rename(columns={"free_qty": "actual_total"}))
    one = inv.drop_duplicates(subset=["row_key"], keep="first").merge(agg, on="row_key")

    rows: list[dict] = []
    for _, r in one.iterrows():
        key = str(r["row_key"])
        actual = float(r["actual_total"])
        led = exact.get(key) or soft.get(
            _soft_key(r["pallet"], r["item_number"], r["lot_number"]))

        if not led:
            mode, avail = "NEW", actual
            lb = lp = lbal = 0.0
        else:
            lb, lp, lbal = led["before"], led["picked"], led["balance"]
            if abs(actual - lb) <= QTY_TOL:
                mode, avail = "LEDGER BALANCE", lbal      # inventory එක පරණයි
            else:
                mode, avail = "NEW BASELINE", actual      # inventory update වෙලා

        free = bool(r.get("pick_free", True))
        rows.append({
            "PALLET": r["pallet"], "LOCATION_ID": r["location_id"],
            "ITEM_NUMBER": r["item_number_raw"] or r["item_number"],
            "BASE_ID": r["base_id"], "LOT_NUMBER": r["lot_number"], "PLANT": r["plant"],
            "UOM": r["uom"], "PICK_ID": r.get("pick_id", ""),
            "PICK_STATUS": "FREE" if free else "ON PICK TASK",
            "INV_ACTUAL_QTY": actual,
            "LEDGER_BEFORE": lb, "LEDGER_PICKED": lp, "LEDGER_BALANCE": lbal,
            "MODE": mode, "AVAILABLE": max(0.0, float(avail)), "ROW_KEY": key,
        })

    basis = pd.DataFrame(rows, columns=BASIS_COLS)
    avail_map = dict(zip(basis["ROW_KEY"], basis["AVAILABLE"]))
    one = one.copy()
    one["avail"] = one["row_key"].map(avail_map).fillna(0.0)
    return one.reset_index(drop=True), basis


# --------------------------------------------------------------------------- #
# Config
# --------------------------------------------------------------------------- #
@dataclass
class EngineConfig:
    wh_id: str = "INMM01"
    client_code: str = "INM0DONA"
    order_type: str = "Sales Orders"
    plants: list[str] = field(default_factory=list)
    statuses: list[str] = field(default_factory=lambda: ["Available"])
    pick_id_zero_only: bool = True       # Pick Id 0 විතරක් — 0 නොවන එක locked
    # user confirm කරලා release කරපු ඒවා: {doc number: [pick id, …]}  ("*" = ඔක්කොම)
    release_locked: dict[str, list[str]] = field(default_factory=dict)
    # Partial pick — තියෙන ප්‍රමාණය විතරක් යවන්න user confirm කරපු documents.
    # ["*"] = ඔක්කොම. හිස් නම් කලින් වගේම all-or-nothing.
    partial_docs: list[str] = field(default_factory=list)
    # How much of a short line a partial pick takes.
    #   "floor" — whatever is on the floor, so a line can go out split
    #   "whole" — only lines that can be filled completely; a short line is
    #             left out of the load entirely and stays owed in full
    partial_mode: str = "floor"
    # Lines the user has deliberately left out of this load: {doc: [line no, …]}.
    # The quantity stays owed, so the document comes back for it.
    skip_lines: dict[str, list[int]] = field(default_factory=dict)
    strategy: str = "FIFO"              # FIFO | LEAST_PALLETS | SINGLE_PALLET_FIRST
    exact_item_first: bool = True
    use_ledger: bool = True             # pallet balance එකට ledger එක බලනවද
    blank_fill: str = "TBC"
    fill_item_number_col: bool = False
    fill_description: bool = False       # OutBound Detail එකේ ITEM_DESCRIPTION පුරවනවද
    merge_same_item_lines: bool = False
    override_doc_check: bool = False    # ⚠️ manual verify කරලා විතරක් — stock check bypass වෙන්නේ නෑ
    pick_date: datetime = field(default_factory=datetime.now)


# --------------------------------------------------------------------------- #
# Allocation
# --------------------------------------------------------------------------- #
def _order_pool(pool: pd.DataFrame, need: float, doc_item: str, cfg: EngineConfig) -> pd.DataFrame:
    p = pool.copy()
    p["_exact"] = (p["item_number"] == clean_item(doc_item)).astype(int)
    asc_exact = not cfg.exact_item_first

    if cfg.strategy == "LEAST_PALLETS":
        p = p.sort_values(["_exact", "avail", "_fifo"], ascending=[asc_exact, False, True])
    elif cfg.strategy == "SINGLE_PALLET_FIRST":
        p["_covers"] = (p["avail"] >= need).astype(int)
        p = p.sort_values(["_exact", "_covers", "avail", "_fifo"],
                          ascending=[asc_exact, False, True, True])
    else:                                                    # FIFO
        p = p.sort_values(["_exact", "_fifo", "avail"], ascending=[asc_exact, True, True])
    return p


def _detail_key(a: dict) -> tuple:
    return (a["ITEM_NUMBER"], a["LOT_NUMBER"], a["UOM"], a["PLANT"],
            a["_attrs"])


def run_pick(
    docs: list[ParsedDoc],
    inv_raw: pd.DataFrame,
    cfg: EngineConfig,
    ledger: pd.DataFrame | None = None,
    processed_docs: set[str] | None = None,
    sku_desc: dict[str, str] | None = None,
    picked_before: dict[str, dict[int, float]] | None = None,
) -> dict[str, Any]:
    """
    මුළු pipeline එක — validate -> duplicate -> allocate -> WMS output.

    `picked_before` = {doc number: {line no: qty}} — කලින් run එකක partial
    විදිහට pick කරපු ප්‍රමාණය. ඒ ටික අඩු කරලා **ඉතුරු ටික විතරයි** මේ run
    එකේ pick වෙන්නේ.
    """
    run_id = datetime.now().strftime("%Y%m%d%H%M%S") + "-" + uuid.uuid4().hex[:4]
    stamp = cfg.pick_date.strftime("%Y-%m-%d %H:%M:%S")
    processed_docs = {str(x).strip() for x in (processed_docs or set())}
    sku_desc = sku_desc or {}
    prev_map = {str(k).strip(): {int(l): float(q) for l, q in (v or {}).items()}
                for k, v in (picked_before or {}).items()}
    partial_ok = {str(x).strip() for x in (cfg.partial_docs or [])}

    def _desc(inv_d: Any, item: Any, doc_d: Any = "") -> str:
        """Description — inventory -> SKU master -> document."""
        for v in (inv_d, sku_desc.get(clean_item(item)), sku_desc.get(base_item(item)),
                  doc_d):
            if str(v or "").strip():
                return str(v).strip()
        return ""

    inv = normalize_inventory(inv_raw)
    if cfg.statuses:
        inv = inv[inv["status"].isin(cfg.statuses)]
    if cfg.plants:
        inv = inv[inv["plant"].isin(cfg.plants)]

    # ---- Pick Id — 0 නොවන pallet දැනටමත් pick task එකකට allocate වෙලා ----
    # Locked rows stay in the pool so a released document can still reach them;
    # the gate is applied per document inside the loop.
    released: dict[str, set[str]] = {str(k).strip(): {str(v).strip() for v in (ids or [])}
                                     for k, ids in (cfg.release_locked or {}).items()}
    locked = inv[~inv["pick_free"]].copy() if cfg.pick_id_zero_only else inv.iloc[0:0].copy()

    # ---- pallet level stock basis (QTY_BEFORE same ද කියලා බලලා) ----
    inv, basis = stock_basis(inv, ledger, use_ledger=cfg.use_ledger)
    remaining: dict[str, float] = dict(zip(basis["ROW_KEY"], basis["AVAILABLE"].astype(float)))
    cap: dict[str, float] = dict(remaining)          # pallet එකට වඩා pick වෙන්න බෑ
    inv = inv[inv["avail"] > 0].reset_index(drop=True)

    allocations: list[dict] = []
    verify_rows: list[dict] = []
    rejected: list[dict] = []
    shortages: list[dict] = []
    accepted: list[dict] = []
    partials: list[str] = []
    offers: list[dict] = []
    detail_rows: list[dict] = []
    master_rows: list[dict] = []

    seen_batch: set[str] = set()

    for doc in docs:
        num = str(doc.doc_number).strip()

        # ---------- duplicate ----------
        if num in seen_batch:
            rejected.append({"DOC_NUMBER": num, "DOC_TYPE": doc.doc_type,
                             "REASON": "DUPLICATE (batch)",
                             "DETAIL": "Same number appears twice in this upload - kept the first one",
                             "SOURCE_FILE": doc.source_file})
            continue
        if num in processed_docs:
            rejected.append({"DOC_NUMBER": num, "DOC_TYPE": doc.doc_type,
                             "REASON": "DUPLICATE (already processed)",
                             "DETAIL": "Already picked in an earlier run",
                             "SOURCE_FILE": doc.source_file})
            continue
        seen_batch.add(num)

        # ---------- document completeness ----------
        ok, problems = doc.completeness()
        doc_check = "OK"
        if not ok:
            if not cfg.override_doc_check:
                rejected.append({"DOC_NUMBER": num, "DOC_TYPE": doc.doc_type,
                                 "REASON": "INCOMPLETE DOCUMENT",
                                 "DETAIL": " · ".join(problems),
                                 "SOURCE_FILE": doc.source_file})
                continue
            doc_check = "MANUAL OVERRIDE — " + " · ".join(problems)

        # ---------- dry-run allocation (all-or-nothing) ----------
        trial = {k: v for k, v in remaining.items()}
        doc_alloc: list[dict] = []
        doc_short: list[dict] = []
        prev = prev_map.get(num, {})
        go_partial = bool(partial_ok) and ("*" in partial_ok or num in partial_ok)
        allow = released.get(num, set())
        all_released = "*" in allow
        used_ids: set[str] = set()

        def _open(df: pd.DataFrame) -> pd.Series:
            """Rows this document may take from — free, or released by the user."""
            if not cfg.pick_id_zero_only:
                return pd.Series(True, index=df.index)
            ok = df["pick_free"]
            if all_released:
                return pd.Series(True, index=df.index)
            if allow:
                ok = ok | df["pick_id"].isin(allow)
            return ok

        skip_set = {int(x) for x in (cfg.skip_lines or {}).get(num, []) if str(x).strip()}
        left_out: list[dict] = []

        for ln in doc.lines:
            # only the balance is still owed when an earlier run took some
            done_before = float(prev.get(ln.line_no, 0.0))
            need = float(ln.qty) - done_before
            if need <= QTY_TOL:
                continue                       # this line was completed earlier
            if ln.line_no in skip_set:
                # deliberately held back — not a shortage, a decision. The
                # quantity stays owed so the document comes back for it.
                left_out.append({
                    "DOC_NUMBER": num, "DOC_TYPE": doc.doc_type, "DOC_LINE": ln.line_no,
                    "DOC_ITEM_CODE": ln.item_code, "BASE_ID": ln.base,
                    "DESCRIPTION": _desc("", ln.item_code, ln.description),
                    "REQUIRED": need, "AVAILABLE": 0.0, "ON_PICK_TASK": 0.0,
                    "PICK_IDS": "", "SHORT": need, "PICKED_NOW": 0.0,
                    "REASON": "Left out of this load by the user",
                })
                continue
            # Base id first, but the *exact* code always matches itself too:
            # "P601560710" on the document and "P601560 710" in the inventory
            # clean to the same key even though only one of them can be split
            # into a base. Without this the line reads "Item not in inventory".
            key = clean_item(ln.item_code)
            base_pool = inv[(inv["base_id"] == ln.base)
                            | (inv["item_number"] == key)].copy()
            base_pool["avail"] = base_pool["row_key"].map(trial).fillna(0.0)
            base_pool = base_pool[base_pool["avail"] > 0]
            open_mask = _open(base_pool)
            pool = base_pool[open_mask].copy()
            shut = base_pool[~open_mask]
            have = float(pool["avail"].sum())

            if pool.empty or have + 1e-9 < need:
                lock_q = float(shut["avail"].sum()) if len(shut) else 0.0
                lock_id = (", ".join(sorted({str(x) for x in shut["pick_id"]})[:4])
                           if len(shut) else "")
                if pool.empty and not lock_q:
                    why = "Item not in inventory / plant"
                elif lock_q and have + lock_q + 1e-9 >= need:
                    why = (f"On another pick task — {_qty_str(lock_q)} locked "
                           f"(Pick Id {lock_id})")
                elif lock_q:
                    why = (f"Stock short · {_qty_str(lock_q)} also locked to a pick task "
                           f"(Pick Id {lock_id})")
                else:
                    why = "Stock short"
                # Whole-line mode sends nothing for a short line, so the owed
                # quantity is the whole line — not the balance of a split.
                takes_floor = (go_partial and have > QTY_TOL
                               and str(cfg.partial_mode).lower() != "whole")
                sending = have if takes_floor else 0.0
                doc_short.append({
                    "DOC_NUMBER": num, "DOC_TYPE": doc.doc_type, "DOC_LINE": ln.line_no,
                    "DOC_ITEM_CODE": ln.item_code, "BASE_ID": ln.base,
                    "DESCRIPTION": _desc("", ln.item_code, ln.description),
                    "REQUIRED": need, "AVAILABLE": have,
                    "ON_PICK_TASK": lock_q, "PICK_IDS": lock_id,
                    "SHORT": max(0.0, need - sending), "REASON": why,
                    "PICKED_NOW": sending,
                })
                # A partial pick takes whatever is on the floor and leaves the
                # rest owed; without the user's confirmation the line is simply
                # not picked and the whole document falls over below.
                # In "whole" mode a short line never goes out split — the load
                # carries only the lines that can be filled completely.
                if not takes_floor:
                    continue
                need = have

            for _, r in _order_pool(pool, need, ln.item_code, cfg).iterrows():
                if need <= 1e-9:
                    break
                take = min(need, float(trial.get(r["row_key"], 0.0)))
                if take <= 0:
                    continue
                before = float(trial.get(r["row_key"], 0.0))
                trial[r["row_key"]] = before - take
                need -= take
                if not bool(r.get("pick_free", True)):
                    used_ids.add(str(r.get("pick_id", "")))
                doc_alloc.append({
                    "RUN_ID": run_id, "PICK_DATE": stamp, "DOC_TYPE": doc.doc_type,
                    "DOC_NUMBER": num, "DOC_LINE": ln.line_no,
                    "DOC_ITEM_CODE": ln.item_code, "BASE_ID": ln.base,
                    "ITEM_NUMBER": r["item_number_raw"] or r["item_number"],
                    "DESCRIPTION": _desc(r["description"], r["item_number_raw"],
                                         ln.description),
                    "PALLET": r["pallet"], "LOCATION_ID": r["location_id"],
                    "LOT_NUMBER": r["lot_number"], "PLANT": r["plant"],
                    "UOM": r["uom"] or ln.uom, "QTY_BEFORE": before,
                    "QTY_PICKED": take, "QTY_BALANCE": before - take,
                    "FIFO_DATE": r["fifo_date"], "GRN_NUMBER": r["grn_number"],
                    "STORED_ATTRIBUTE_ID": r["stored_attribute_id"],
                    "ROW_KEY": r["row_key"], "SOURCE_FILE": doc.source_file,
                    "_attrs": tuple(str(r.get(v, "") or "") for v in GEN_MAP.values()),
                })

        # A line the user held back is owed just like a short one, and asking
        # for it is itself a request for a partial pick.
        if left_out:
            doc_short.extend(left_out)
            go_partial = True

        # ---------- whole document must be complete ----------
        # …unless the user has confirmed a partial pick for this document, and
        # there is actually something to send.
        miss = ", ".join(
            f"L{s['DOC_LINE']} {s['DOC_ITEM_CODE']} (need {s['REQUIRED']:g}, "
            f"have {s['AVAILABLE']:g})" for s in doc_short
        )
        if doc_short:
            shortages.extend(doc_short)
            if not (go_partial and doc_alloc):
                rejected.append({"DOC_NUMBER": num, "DOC_TYPE": doc.doc_type,
                                 "REASON": "STOCK SHORT - not picked",
                                 "DETAIL": miss, "SOURCE_FILE": doc.source_file})
                # What a partial pick would actually put on the truck, worked
                # out here because this is the only place that knows it: the
                # lines that were fine allocated in full (doc_alloc), and each
                # short line could still give whatever is on the floor.
                # Deriving it from the shortage table alone was wrong — that
                # table holds only the short lines, so a document with one dead
                # line and five good ones looked like it had nothing to send.
                whole_only = sum(float(a["QTY_PICKED"]) for a in doc_alloc)
                can_now = whole_only + sum(float(x["AVAILABLE"]) for x in doc_short)
                doc_total = float(sum(l.qty for l in doc.lines))
                offers.append({
                    "DOC_NUMBER": num, "DOC_TYPE": doc.doc_type,
                    "LINES": len(doc.lines), "SHORT_LINES": len(doc_short),
                    "DOC_QTY": doc_total,
                    "ALREADY_SENT": float(sum(prev.values())),
                    "CAN_PICK_NOW": can_now,
                    # what would go out if short lines were left whole instead
                    # of split — the lines that are complete, and nothing else
                    "WHOLE_LINES_ONLY": whole_only,
                    "COMPLETE_LINES": len(doc.lines) - len(doc_short),
                    "STILL_SHORT": max(0.0, doc_total - float(sum(prev.values())) - can_now),
                    "ITEMS": ", ".join(dict.fromkeys(
                        str(x["DOC_ITEM_CODE"]) for x in doc_short))[:200],
                    # why it is short, so a document that can send nothing says
                    # so on its own row instead of sending the user hunting
                    "REASONS": ", ".join(dict.fromkeys(
                        str(x["REASON"]) for x in doc_short))[:160],
                })
                continue
        is_partial = bool(doc_short)

        # ---------- WMS Detail (pallet allocations -> order lines) ----------
        groups: dict[tuple, dict] = {}
        order: list[tuple] = []
        for a in doc_alloc:
            key = _detail_key(a) if cfg.merge_same_item_lines else (a["DOC_LINE"],) + _detail_key(a)
            if key not in groups:
                groups[key] = {"qty": 0.0, "a": a}
                order.append(key)
            groups[key]["qty"] += a["QTY_PICKED"]

        doc_detail: list[dict] = []
        line_no = 0
        for key in order:
            g = groups[key]
            a = g["a"]
            line_no += 1
            row = {c: "" for c in DETAIL_COLS}
            row.update(DETAIL_FIXED)
            row["WH_ID"] = cfg.wh_id
            row["CLIENT_CODE"] = cfg.client_code
            row["DISPLAY_ORDER_NUMBER"] = num
            row["LINE_NUMBER"] = str(line_no)
            row["DISPLAY_ITEM_NUMBER"] = a["ITEM_NUMBER"]
            if cfg.fill_description:
                row["ITEM_DESCRIPTION"] = a["DESCRIPTION"]
            if cfg.fill_item_number_col:
                row["ITEM_NUMBER"] = a["ITEM_NUMBER"]
            row["LOT_NUMBER"] = a["LOT_NUMBER"] or cfg.blank_fill
            row["QTY"] = _qty_str(g["qty"])
            row["ORDER_UOM"] = a["UOM"] or cfg.blank_fill
            for i, (gcol, invcol) in enumerate(GEN_MAP.items()):
                val = a["_attrs"][i]
                row[gcol] = val if str(val).strip() else cfg.blank_fill
            doc_detail.append(row)

        # ---------- PALLET CAP — pallet එකට වඩා වැඩියෙන් pick වෙන්න බෑ ----------
        per_key: dict[str, float] = {}
        for a in doc_alloc:
            per_key[a["ROW_KEY"]] = per_key.get(a["ROW_KEY"], 0.0) + float(a["QTY_PICKED"])
        over = [
            f"{k.split('|')[0]} {k.split('|')[2] if len(k.split('|')) > 2 else ''}: "
            f"pick {q:g} > balance {remaining.get(k, 0.0):g}"
            for k, q in per_key.items() if q > remaining.get(k, 0.0) + QTY_TOL
        ]
        if over:
            rejected.append({"DOC_NUMBER": num, "DOC_TYPE": doc.doc_type,
                             "REASON": "PALLET OVER-PICK",
                             "DETAIL": " · ".join(over), "SOURCE_FILE": doc.source_file})
            continue

        # ---------- QUANTITY VERIFY — Invoice / DC qty එකට හරියටම ----------
        v_rows, v_bad = _verify_doc(doc, doc_alloc, doc_detail, prev=prev,
                                    partial=is_partial)
        verify_rows.extend(v_rows)
        if v_bad:
            rejected.append({"DOC_NUMBER": num, "DOC_TYPE": doc.doc_type,
                             "REASON": "QTY VERIFY FAILED",
                             "DETAIL": " · ".join(v_bad), "SOURCE_FILE": doc.source_file})
            continue
        if is_partial:
            partials.append(num)

        # ---------- commit ----------
        remaining = trial
        allocations.extend(doc_alloc)
        detail_rows.extend(doc_detail)

        # ---------- WMS Master ----------
        m = {c: "" for c in MASTER_COLS}
        m.update(MASTER_FIXED)
        m["ORDER_TYPE"] = cfg.order_type
        m["WH_ID"] = cfg.wh_id
        m["CLIENT_CODE"] = cfg.client_code
        m["DISPLAY_ORDER_NUMBER"] = num
        m["STORE_ORDER_NUMBER"] = num
        m["CUSTOMER_PO_NUMBER"] = num
        m["LOAD_ID"] = num
        master_rows.append(m)

        rel_note = ", ".join(sorted(x for x in used_ids if x))
        if rel_note:
            doc_check = (doc_check + " · " if doc_check != "OK" else "") + \
                f"RELEASED from pick task {rel_note}"
        if is_partial:
            doc_check = (doc_check + " · " if doc_check != "OK" else "") + \
                f"PARTIAL PICK — short: {miss}"
        doc_qty = float(sum(l.qty for l in doc.lines))
        now_qty = float(sum(a["QTY_PICKED"] for a in doc_alloc))
        prev_qty = float(sum(prev.values()))
        accepted.append({
            "DOC_NUMBER": num, "LOAD_ID": num, "DOC_TYPE": doc.doc_type,
            "DOC_DATE": doc.doc_date, "RELEASED": rel_note,
            "REF_NUMBER": doc.ref_number, "DOC_CHECK": doc_check, "LINES": len(doc.lines),
            "WMS_LINES": line_no, "DOC_QTY": doc_qty,
            "PICKED_QTY": now_qty,
            "WMS_QTY": sum(float(r["QTY"]) for r in doc_detail),
            # what the document has received in total, this run and before it
            "PREV_QTY": prev_qty, "TOTAL_PICKED": prev_qty + now_qty,
            "SHORT_QTY": max(0.0, doc_qty - prev_qty - now_qty),
            "PICK_STATUS": "PARTIAL" if is_partial else "FULL",
            "VERIFY": "⚠️ PARTIAL" if is_partial else "✅ OK",
            "PALLETS": len({a["PALLET"] for a in doc_alloc}),
            "PLANTS": ", ".join(sorted({a["PLANT"] for a in doc_alloc})),
            "SOURCE_FILE": doc.source_file,
        })

    alloc_df = pd.DataFrame(allocations)
    if len(alloc_df):
        alloc_df = alloc_df[ALLOC_COLS]

    master_df = pd.DataFrame(master_rows, columns=MASTER_COLS)
    detail_df = pd.DataFrame(detail_rows, columns=DETAIL_COLS)

    return {
        "run_id": run_id,
        "pick_date": stamp,
        "master": master_df,
        "detail": detail_df,
        "allocations": alloc_df if len(alloc_df) else pd.DataFrame(columns=ALLOC_COLS),
        "rejected": pd.DataFrame(rejected, columns=["DOC_NUMBER", "DOC_TYPE", "REASON",
                                                    "DETAIL", "SOURCE_FILE"]),
        "shortage": pd.DataFrame(shortages),
        "verify": pd.DataFrame(verify_rows, columns=VERIFY_COLS),
        "basis": basis,
        "locked": _locked_view(locked, released),
        "accepted": pd.DataFrame(accepted),
        "balance": pallet_balance(basis, alloc_df, cap),
        "partial": partials,
        "partial_offer": pd.DataFrame(offers, columns=PARTIAL_COLS),
        "picked_before": prev_map,
        "cfg": cfg,
    }


def _qty_str(q: float) -> str:
    return str(int(round(q))) if abs(q - round(q)) < 1e-9 else f"{q:g}"


# --------------------------------------------------------------------------- #
# Quantity verification — Invoice / DC qty එකට හරියටම ගැලපෙනවද?
# --------------------------------------------------------------------------- #
VERIFY_COLS = ["DOC_NUMBER", "DOC_TYPE", "LINE", "ITEM_CODE", "ITEM_NUMBER",
               "DOC_QTY", "PICKED_QTY", "WMS_QTY", "DIFF", "STATUS"]

_TOL = 1e-6


def _verify_doc(doc, doc_alloc: list[dict], doc_detail: list[dict],
                prev: dict[int, float] | None = None,
                partial: bool = False) -> tuple[list[dict], list[str]]:
    """
    Line by line + document total + WMS file total — තුනම match වෙන්න ඕන.
    Fail වුණොත් document එක reject වෙනවා.

    Full pick එකකදී line එකකට **owed** = doc qty − කලින් pick කරපු ප්‍රමාණය,
    ඒක හරියටම pick වෙන්න ඕන. Partial pick එකකදී අඩුවෙන් යවන එක **තීරණයක්**,
    ඒ නිසා check කරන්නේ "WMS file එකට ගියේ ඇත්තටම pallet එකෙන් අරන් තියෙන
    ප්‍රමාණයමද, ඕන ප්‍රමාණයට වඩා වැඩිද" කියන එක.
    """
    prev = prev or {}
    picked: dict[int, float] = {}
    items: dict[int, set[str]] = {}
    for a in doc_alloc:
        picked[a["DOC_LINE"]] = picked.get(a["DOC_LINE"], 0.0) + float(a["QTY_PICKED"])
        items.setdefault(a["DOC_LINE"], set()).add(str(a["ITEM_NUMBER"]))

    rows: list[dict] = []
    bad: list[str] = []

    for ln in doc.lines:
        p = picked.get(ln.line_no, 0.0)
        before = float(prev.get(ln.line_no, 0.0))
        owed = float(ln.qty) - before
        diff = p - owed
        if partial:
            # under is allowed and expected; over never is
            ok = diff <= _TOL
            status = "✅ OK" if abs(diff) <= _TOL else "⚠️ SHORT"
            if not ok:
                bad.append(f"L{ln.line_no} {ln.item_code}: picked {p:g} > owed {owed:g}")
        else:
            ok = abs(diff) <= _TOL
            status = "✅ OK" if ok else "❌ MISMATCH"
            if not ok:
                bad.append(f"L{ln.line_no} {ln.item_code}: doc {owed:g} ≠ picked {p:g}")
        rows.append({
            "DOC_NUMBER": doc.doc_number, "DOC_TYPE": doc.doc_type,
            "LINE": str(ln.line_no), "ITEM_CODE": ln.item_code,
            "ITEM_NUMBER": ", ".join(sorted(items.get(ln.line_no, set()))),
            "DOC_QTY": float(ln.qty), "PICKED_QTY": p + before,
            "WMS_QTY": p, "DIFF": (p + before) - float(ln.qty), "STATUS": status,
        })

    doc_total = float(sum(l.qty for l in doc.lines))
    prev_total = float(sum(prev.get(l.line_no, 0.0) for l in doc.lines))
    owed_total = doc_total - prev_total
    pick_total = float(sum(picked.values()))
    wms_total = float(sum(float(r["QTY"]) for r in doc_detail))

    # This one holds either way: the WMS file must carry exactly what came off
    # the pallets, no more and no less.
    if abs(wms_total - pick_total) > _TOL:
        bad.append(f"WMS file total: picked {pick_total:g} ≠ OutBound Detail {wms_total:g}")
    if partial:
        if pick_total > owed_total + _TOL:
            bad.append(f"Document total: picked {pick_total:g} > owed {owed_total:g}")
    else:
        if abs(pick_total - owed_total) > _TOL:
            bad.append(f"Document total: doc {owed_total:g} ≠ picked {pick_total:g}")
        # the document's own "Total Quantity" only speaks for the whole document
        if not prev and doc.declared_qty is not None \
                and abs(doc.declared_qty - pick_total) > _TOL:
            bad.append(f"Document 'Total Quantity' {doc.declared_qty:g} "
                       f"≠ picked {pick_total:g}")

    rows.append({
        "DOC_NUMBER": doc.doc_number, "DOC_TYPE": doc.doc_type, "LINE": "TOTAL",
        "ITEM_CODE": f"{len(doc.lines)} lines",
        "ITEM_NUMBER": (f"doc says {doc.declared_qty:g}"
                        if doc.declared_qty is not None else ""),
        "DOC_QTY": doc_total, "PICKED_QTY": pick_total + prev_total, "WMS_QTY": wms_total,
        "DIFF": (pick_total + prev_total) - doc_total,
        "STATUS": ("❌ MISMATCH" if bad else
                   ("⚠️ PARTIAL" if partial else "✅ OK")),
    })
    return rows, bad


PARTIAL_COLS = ["DOC_NUMBER", "DOC_TYPE", "LINES", "COMPLETE_LINES", "SHORT_LINES",
                "DOC_QTY", "ALREADY_SENT", "CAN_PICK_NOW", "WHOLE_LINES_ONLY",
                "STILL_SHORT", "ITEMS", "REASONS"]


def partialable(res: dict[str, Any]) -> pd.DataFrame:
    """
    Documents refused for stock that could still put something on a truck —
    the answer to "we cannot wait, send what we have".

    A document with nothing at all in the warehouse is not a partial pick, it
    is no pick, so it is left out; `no_partial()` names those separately rather
    than leaving the user hunting for a section that never appears.
    """
    out = res.get("partial_offer")
    if out is None or not len(out):
        return pd.DataFrame(columns=PARTIAL_COLS)
    room = out["CAN_PICK_NOW"].astype(float)
    if "WHOLE_LINES_ONLY" in out.columns:
        room = room.combine(out["WHOLE_LINES_ONLY"].astype(float), max)
    return out[room > 0].reset_index(drop=True)


def no_partial(res: dict[str, Any]) -> pd.DataFrame:
    """Stock-short documents where a partial pick would send nothing at all."""
    out = res.get("partial_offer")
    if out is None or not len(out):
        return pd.DataFrame(columns=PARTIAL_COLS)
    room = out["CAN_PICK_NOW"].astype(float)
    if "WHOLE_LINES_ONLY" in out.columns:
        room = room.combine(out["WHOLE_LINES_ONLY"].astype(float), max)
    return out[room <= 0].reset_index(drop=True)


RELEASE_COLS = ["DOC_NUMBER", "DOC_TYPE", "SHORT_LINES", "SHORT_QTY", "ON_PICK_TASK",
                "PICK_IDS", "ITEMS"]


def releasable(res: dict[str, Any]) -> pd.DataFrame:
    """
    Documents that failed **only** because the stock sits on another pick task —
    every short line would be covered if those pallets were released.
    Anything genuinely out of stock is not listed: a confirmation cannot invent
    inventory.
    """
    sh = res.get("shortage")
    if sh is None or not len(sh) or "ON_PICK_TASK" not in sh.columns:
        return pd.DataFrame(columns=RELEASE_COLS)

    d = sh.copy()
    for c in ("REQUIRED", "AVAILABLE", "ON_PICK_TASK", "SHORT"):
        d[c] = pd.to_numeric(d.get(c), errors="coerce").fillna(0.0)

    rows: list[dict] = []
    for num, g in d.groupby("DOC_NUMBER"):
        if not (g["ON_PICK_TASK"] > 0).any():
            continue
        if not ((g["AVAILABLE"] + g["ON_PICK_TASK"]) + QTY_TOL >= g["REQUIRED"]).all():
            continue                      # some line is short even after a release
        # The reason string caps the ids for readability — the release set must
        # be complete, so take it from the locked frame by base id instead.
        bases = {str(b) for b in g["BASE_ID"]}
        lk = res.get("locked")
        ids: set[str] = set()
        if lk is not None and len(lk) and "BASE_ID" in lk.columns:
            ids = {str(x) for x in lk.loc[lk["BASE_ID"].astype(str).isin(bases), "PICK_ID"]
                   if str(x).strip()}
        if not ids:
            for v in g.get("PICK_IDS", pd.Series(dtype=str)):
                ids |= {x.strip() for x in str(v).split(",") if x.strip()}
        rows.append({
            "DOC_NUMBER": str(num), "DOC_TYPE": str(g["DOC_TYPE"].iloc[0]),
            "SHORT_LINES": int(len(g)), "SHORT_QTY": float(g["SHORT"].sum()),
            "ON_PICK_TASK": float(g["ON_PICK_TASK"].sum()),
            "PICK_IDS": ", ".join(sorted(ids)),
            "ITEMS": ", ".join(sorted({str(x) for x in g["DOC_ITEM_CODE"]})),
        })
    return pd.DataFrame(rows, columns=RELEASE_COLS)


# --------------------------------------------------------------------------- #
# Per LOAD_ID slicing / files
# --------------------------------------------------------------------------- #
def safe_name(text: Any, fallback: str = "DOC") -> str:
    """'333/26-27/62' -> '333-26-27-62' (filename safe LOAD_ID)"""
    s = re.sub(r"[\\/:*?\"<>|\s]+", "-", str(text or "").strip())
    s = re.sub(r"-{2,}", "-", s).strip("-.")
    return s or fallback


def _slice(df: pd.DataFrame, col: str, value: str) -> pd.DataFrame:
    if df is None or not len(df) or col not in df.columns:
        return pd.DataFrame(columns=df.columns if df is not None else [])
    return df[df[col].astype(str) == str(value)].reset_index(drop=True)


def doc_bundle(res: dict[str, Any], load_id: str) -> dict[str, Any]:
    """එක LOAD_ID එකකට අදාල ඔක්කොම — master · detail · allocation · verify · info."""
    acc = res["accepted"]
    row = acc[acc["DOC_NUMBER"].astype(str) == str(load_id)]
    info: dict[str, Any] = row.iloc[0].to_dict() if len(row) else {"DOC_NUMBER": load_id}
    cfg: EngineConfig = res.get("cfg") or EngineConfig()
    master = _slice(res["master"], "DISPLAY_ORDER_NUMBER", load_id)
    detail = _slice(res["detail"], "DISPLAY_ORDER_NUMBER", load_id)
    alloc = _slice(res["allocations"], "DOC_NUMBER", load_id)
    verify = _slice(res.get("verify", pd.DataFrame()), "DOC_NUMBER", load_id)
    return {
        "load_id": str(load_id),
        "safe": safe_name(load_id),
        "master": master, "detail": detail, "allocations": alloc, "verify": verify,
        "info": {
            "LOAD_ID": str(load_id),
            "DOC_NUMBER": str(load_id),
            "DOC_TYPE": info.get("DOC_TYPE", ""),
            "DOC_DATE": info.get("DOC_DATE", ""),
            "REF_NUMBER": info.get("REF_NUMBER", ""),
            "PLANT": info.get("PLANTS", ""),
            "LINES": info.get("LINES", len(detail)),
            "TOTAL_QTY": info.get("DOC_QTY", 0),
            "PALLETS": info.get("PALLETS", 0),
            "VERIFY": info.get("VERIFY", ""),
            "RELEASED": info.get("RELEASED", ""),
            "PICK_STATUS": info.get("PICK_STATUS", "FULL"),
            "PICKED_QTY": info.get("PICKED_QTY", 0),
            "PREV_QTY": info.get("PREV_QTY", 0),
            "SHORT_QTY": info.get("SHORT_QTY", 0),
            "SOURCE_FILE": info.get("SOURCE_FILE", ""),
            "RUN_ID": res.get("run_id", ""),
            "PICK_DATE": res.get("pick_date", ""),
            "STRATEGY": cfg.strategy,
            "WH_ID": cfg.wh_id,
            "CLIENT": cfg.client_code,
        },
    }


def load_ids(res: dict[str, Any]) -> list[str]:
    acc = res.get("accepted")
    if acc is None or not len(acc):
        return []
    return [str(x) for x in acc["DOC_NUMBER"].tolist()]


def build_zip(files: list[tuple[str, bytes]]) -> bytes:
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in files:
            if data:
                z.writestr(name, data)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Global search
# --------------------------------------------------------------------------- #
def search_frames(query: str, frames: dict[str, pd.DataFrame],
                  limit: int = 400) -> dict[str, pd.DataFrame]:
    """
    ඕනෑම data එකක් — word කීපයක් දුන්නොත් ඔක්කොම තියෙන rows විතරයි (AND).
    """
    terms = [t.strip().lower() for t in str(query).split() if t.strip()]
    out: dict[str, pd.DataFrame] = {}
    if not terms:
        return out
    for name, df in frames.items():
        if df is None or not len(df):
            continue
        try:
            joined = None
            for c in df.columns:
                col = df[c].astype("string").fillna("").str.lower()
                joined = col if joined is None else joined.str.cat(col, sep=" | ")
            if joined is None:
                continue
            mask = pd.Series(True, index=df.index)
            for t in terms:
                mask &= joined.str.contains(re.escape(t), na=False)
            hit = df[mask]
            if len(hit):
                out[name] = hit.head(limit).reset_index(drop=True)
        except Exception:
            continue
    return out


def pallet_balance(basis: pd.DataFrame, alloc_df: pd.DataFrame,
                   cap: dict[str, float] | None = None) -> pd.DataFrame:
    """මේ run එකේ pick කරපු හැම pallet එකකම — before · picked · balance."""
    cols = ["PALLET", "LOCATION_ID", "ITEM_NUMBER", "LOT_NUMBER", "PLANT", "UOM",
            "MODE", "INV_ACTUAL_QTY", "QTY_BEFORE", "QTY_PICKED", "QTY_BALANCE", "ROW_KEY"]
    if alloc_df is None or not len(alloc_df):
        return pd.DataFrame(columns=cols)

    g = (alloc_df.groupby("ROW_KEY", dropna=False)
         .agg(QTY_PICKED=("QTY_PICKED", "sum")).reset_index())
    b = basis.set_index("ROW_KEY") if (basis is not None and len(basis)) else None
    out: list[dict] = []
    for _, r in g.iterrows():
        key = str(r["ROW_KEY"])
        info = b.loc[key].to_dict() if (b is not None and key in b.index) else {}
        before = float((cap or {}).get(key, info.get("AVAILABLE", 0.0) or 0.0))
        picked = float(r["QTY_PICKED"])
        out.append({
            "PALLET": info.get("PALLET", key.split("|")[0]),
            "LOCATION_ID": info.get("LOCATION_ID", ""),
            "ITEM_NUMBER": info.get("ITEM_NUMBER", ""),
            "LOT_NUMBER": info.get("LOT_NUMBER", ""), "PLANT": info.get("PLANT", ""),
            "UOM": info.get("UOM", ""), "MODE": info.get("MODE", ""),
            "INV_ACTUAL_QTY": info.get("INV_ACTUAL_QTY", ""),
            "QTY_BEFORE": before, "QTY_PICKED": picked,
            "QTY_BALANCE": before - picked, "ROW_KEY": key,
        })
    return pd.DataFrame(out, columns=cols).sort_values("PALLET").reset_index(drop=True)


LOCKED_COLS = ["PALLET", "LOCATION_ID", "ITEM_NUMBER", "BASE_ID", "DESCRIPTION",
               "LOT_NUMBER", "PLANT", "UOM", "QTY", "PICK_ID", "RELEASED", "STATUS"]


def _locked_view(locked: pd.DataFrame,
                 released: dict[str, set[str]] | None = None) -> pd.DataFrame:
    """Pick Id 0 නොවන නිසා අයින් කරපු stock — කොහෙද, කීයද, මොන pick task එකේද."""
    if locked is None or not len(locked):
        return pd.DataFrame(columns=LOCKED_COLS)
    free_ids: set[str] = set()
    for ids in (released or {}).values():
        free_ids |= {str(i) for i in ids}
    d = locked.rename(columns={
        "pallet": "PALLET", "location_id": "LOCATION_ID", "base_id": "BASE_ID",
        "description": "DESCRIPTION", "lot_number": "LOT_NUMBER", "plant": "PLANT",
        "uom": "UOM", "free_qty": "QTY", "pick_id": "PICK_ID", "status": "STATUS"}).copy()
    d["ITEM_NUMBER"] = d["item_number_raw"].where(d["item_number_raw"].astype(bool),
                                                  d["item_number"])
    d["RELEASED"] = ["YES" if ("*" in free_ids or str(p) in free_ids) else ""
                     for p in d["PICK_ID"]]
    return (d[LOCKED_COLS].sort_values(["BASE_ID", "PALLET"]).reset_index(drop=True))


def stock_view(inv_raw: pd.DataFrame, ledger: pd.DataFrame | None = None,
               use_ledger: bool = True) -> pd.DataFrame:
    """
    Pallet-level live view — Inventory Actual · ledger before/picked ·
    pick කරන්න පුළුවන් BALANCE (QTY_BEFORE same ද කියන rule එකට).
    """
    inv = normalize_inventory(inv_raw)
    _, basis = stock_basis(inv, ledger, use_ledger=use_ledger)
    desc = inv.drop_duplicates("row_key").set_index("row_key")["description"].to_dict()
    b = basis.copy()
    b["DESCRIPTION"] = b["ROW_KEY"].map(desc).fillna("")
    b = b.rename(columns={"INV_ACTUAL_QTY": "ACTUAL_QTY", "LEDGER_PICKED": "PICKED_BEFORE",
                          "AVAILABLE": "BALANCE"})
    b.loc[b["PICK_STATUS"] != "FREE", "BALANCE"] = 0.0     # locked = pick කරන්න බෑ
    return b[["PALLET", "LOCATION_ID", "ITEM_NUMBER", "BASE_ID", "DESCRIPTION",
              "LOT_NUMBER", "PLANT", "UOM", "PICK_ID", "PICK_STATUS", "ACTUAL_QTY",
              "LEDGER_BEFORE", "PICKED_BEFORE", "LEDGER_BALANCE", "MODE", "BALANCE",
              "ROW_KEY"]]


# --------------------------------------------------------------------------- #
# Excel writers (everything as TEXT)
# --------------------------------------------------------------------------- #
def _write_text_sheet(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for _, r in df.iterrows():
        ws.append([None if (pd.isna(v) or str(v) == "") else str(v) for v in r.tolist()])
    for col in range(1, len(df.columns) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = max(12, min(32, len(str(df.columns[col - 1])) + 3))
        for cell in ws[letter]:
            cell.number_format = "@"


def build_wms_excel(master: pd.DataFrame, detail: pd.DataFrame) -> bytes:
    """WMS upload file — 'OutBound MASTER' + 'OutBound Detail', සියල්ල text."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "OutBound MASTER"
    _write_text_sheet(ws1, master)
    ws2 = wb.create_sheet("OutBound Detail")
    _write_text_sheet(ws2, detail)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# xlsxwriter is markedly faster and produces much smaller files; it is
# optional, so fall back to openpyxl when it is not installed.
try:                                     # noqa: SIM105
    import xlsxwriter as _xlsxwriter     # noqa: F401
    _XL_ENGINE = "xlsxwriter"
except Exception:                        # pragma: no cover
    _XL_ENGINE = "openpyxl"


def build_report_excel(res: dict[str, Any]) -> bytes:
    """Pick report — allocations, pallet balance, shortage, rejected."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine=_XL_ENGINE) as xw:
        for name, key in [("Doc Summary", "accepted"), ("Qty Verification", "verify"),
                          ("Pallet Allocation", "allocations"),
                          ("Pallet Balance", "balance"), ("Stock Basis", "basis"),
                          ("Shortage", "shortage"), ("On Pick Task", "locked"),
                          ("Rejected Docs", "rejected")]:
            df = res.get(key)
            if key == "basis" and df is not None and len(df):
                df = df[df["MODE"] != "NEW"]          # ledger history තියෙන ඒවා විතරයි
            if df is None or not len(df):
                df = pd.DataFrame({"info": ["- nothing to show -"]})
            df.to_excel(xw, sheet_name=name[:31], index=False)
    return buf.getvalue()
