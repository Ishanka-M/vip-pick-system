"""
Validation harness for the VIP pick engine.

Pick logic:
    mult       = HJ / SAP           (pieces per requirement/order unit)
                  LOOSE: SAP == HJ -> mult = 1
                  SET  : SAP divides HJ -> mult = set size
    target_pcs = Req Qty * mult
    available  = sum of Inventory Actual Qty for the item
    pick whole order-units capped by available stock; the short part (if any)
    is reported in 'Cannot Pick'. No carton-splitting.

Run:  python test_engine.py   (or)   python -m pytest test_engine.py
"""
import os
import io
import pandas as pd
import pick_engine as pe
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SAMPLE = os.path.join(HERE, "sample_data")
REQ = os.path.join(SAMPLE, "1781617499524_Requament.xlsx")
SKU = os.path.join(SAMPLE, "1781617499524_SKU_MASTER.xlsx")
INV = os.path.join(SAMPLE, "1781617499523_Inventory_Report.xlsx")


def _mk(req_rows, sku_rows, inv_rows):
    req = pd.DataFrame(req_rows)
    sku = pd.DataFrame(sku_rows)
    inv = pd.DataFrame(inv_rows)
    return pe.run_pipeline(req, sku, inv, pe.EngineConfig())


def test_set_multiplier_example():
    """SET: SAP=1, HJ=2, Req=2 -> pick 4 (even if a box holds 4)."""
    res = _mk(
        {"Material": ["X1"], "Material Description": ["d"], "Req Qty.": [2], "Delivery No": [1]},
        {"Material code": ["X1"], "Material Desc": ["d"], "Catergory": ["SET"], "HJ": [2], "SAP": [1]},
        {"Item Number": ["X1"] * 1, "Actual Qty": [4], "Cbm": [0.1]},
    )
    assert int(res["pick"]["HJ Pcs Qty"].iloc[0]) == 4


def test_loose_equals_req():
    """LOOSE (SAP==HJ): pick exactly Req Qty pieces, even from larger cartons."""
    res = _mk(
        {"Material": ["L1"], "Material Description": ["d"], "Req Qty.": [2], "Delivery No": [1]},
        {"Material code": ["L1"], "Material Desc": ["d"], "Catergory": ["LOOSE"], "HJ": [1], "SAP": [1]},
        {"Item Number": ["L1", "L1"], "Actual Qty": [3, 3], "Cbm": [0.1, 0.1]},  # cartons of 3
    )
    # 2 pieces picked even though inventory is in cartons of 3 (no carton-split)
    assert int(res["pick"]["HJ Pcs Qty"].iloc[0]) == 2
    assert len(res["exceptions"]) == 0


def test_set_picks_whole_set():
    """SET SAP=1 HJ=3, Req=2 -> 6 pcs when stock plenty."""
    res = _mk(
        {"Material": ["S1"], "Material Description": ["d"], "Req Qty.": [2], "Delivery No": [1]},
        {"Material code": ["S1"], "Material Desc": ["d"], "Catergory": ["SET"], "HJ": [3], "SAP": [1]},
        {"Item Number": ["S1"] * 10, "Actual Qty": [3] * 10, "Cbm": [0.1] * 10},
    )
    assert int(res["pick"]["HJ Pcs Qty"].iloc[0]) == 6


def test_short_stock():
    """target exceeds available -> pick whole units that fit, report short variance."""
    res = _mk(
        {"Material": ["S2"], "Material Description": ["d"], "Req Qty.": [5], "Delivery No": [1]},
        {"Material code": ["S2"], "Material Desc": ["d"], "Catergory": ["SET"], "HJ": [3], "SAP": [1]},
        {"Item Number": ["S2"] * 4, "Actual Qty": [1, 1, 1, 1], "Cbm": [0.1] * 4},  # avail 4
    )
    # target 15, avail 4 -> 1 whole unit of 3 fits -> pick 3, short 12
    assert int(res["pick"]["HJ Pcs Qty"].iloc[0]) == 3
    assert len(res["exceptions"]) == 1
    assert int(res["exceptions"]["Picked Pcs"].iloc[0]) == 3
    assert int(res["exceptions"]["Short Variance"].iloc[0]) == 12


def test_not_in_sku_is_nodata():
    res = _mk(
        {"Material": ["Z9"], "Material Description": ["d"], "Req Qty.": [2], "Delivery No": [1]},
        {"Material code": ["OTHER"], "Material Desc": ["d"], "Catergory": ["LOOSE"], "HJ": [1], "SAP": [1]},
        {"Item Number": ["Z9"], "Actual Qty": [5], "Cbm": [0.1]},
    )
    assert "NO_DATA" in res["full_pick"]["_status"].tolist()
    assert len(res["pick"]) == 0


def test_cannot_exceed_available():
    """Never pick more than the sum of Actual Qty."""
    res = _mk(
        {"Material": ["A1"], "Material Description": ["d"], "Req Qty.": [100], "Delivery No": [1]},
        {"Material code": ["A1"], "Material Desc": ["d"], "Catergory": ["LOOSE"], "HJ": [1], "SAP": [1]},
        {"Item Number": ["A1"] * 3, "Actual Qty": [2, 2, 2], "Cbm": [0.1] * 3},  # avail 6
    )
    assert int(res["pick"]["HJ Pcs Qty"].iloc[0]) <= 6


def test_files_generate_and_format():
    res = _mk(
        {"Material": ["X1"], "Material Description": ["d"], "Req Qty.": [2], "Delivery No": ["110727428A"]},
        {"Material code": ["X1"], "Material Desc": ["d"], "Catergory": ["SET"], "HJ": [2], "SAP": [1]},
        {"Item Number": ["X1", "X1"], "Actual Qty": [2, 2], "Cbm": [0.1, 0.1]},
    )
    assert len(res["vip_pick_bytes"]) > 4000
    assert len(res["india_so_bytes"]) > 4000
    # empty cells must be truly empty (type 'n'), not inline strings (HighJump)
    wb = load_workbook(io.BytesIO(res["india_so_bytes"]))["OutBound MASTER"]
    hdr = [c.value for c in wb[1]]
    cl = wb.cell(row=2, column=hdr.index("CARTON_LABEL") + 1)
    assert cl.data_type == "n", "CARTON_LABEL must be a truly-empty cell"


def test_india_detail_equals_pcs():
    res = _mk(
        {"Material": ["X1", "Y2"], "Material Description": ["d", "d"], "Req Qty.": [2, 3], "Delivery No": [1, 1]},
        {"Material code": ["X1", "Y2"], "Material Desc": ["d", "d"], "Catergory": ["SET", "LOOSE"],
         "HJ": [2, 1], "SAP": [1, 1]},
        {"Item Number": ["X1", "X1", "Y2", "Y2", "Y2"], "Actual Qty": [2, 2, 1, 1, 1], "Cbm": [0.1] * 5},
    )
    detail = res["india_detail"]
    pick = res["pick"]
    qty_col = [c for c in detail.columns if c.upper() == "QTY"][0]
    assert pd.to_numeric(detail[qty_col]).astype(int).tolist() == \
           pd.to_numeric(pick["HJ Pcs Qty"]).astype(int).tolist()


def test_gen_attributes_from_inventory():
    """GEN_ATTRIBUTE_VALUE1..11 pulled from inventory columns."""
    res = _mk(
        {"Material": ["X1"], "Material Description": ["d"], "Req Qty.": [1], "Delivery No": [1]},
        {"Material code": ["X1"], "Material Desc": ["d"], "Catergory": ["LOOSE"], "HJ": [1], "SAP": [1]},
        {"Item Number": ["X1"], "Actual Qty": [5], "Cbm": [0.1],
         "Color": ["RED"], "Size": ["M"], "Style": ["ST1"], "Supplier": ["SUP9"],
         "Plant": ["P1"], "Client So": ["CS1"], "Client So Line": ["10"],
         "Po Cust Dec": ["PC"], "Customer Ref Number": ["CR1"], "Item Id": ["IT1"],
         "Invoice Number1": ["INV1"]},
    )
    d = res["india_detail"].iloc[0]
    assert d["GEN_ATTRIBUTE_VALUE1"] == "RED"     # Color
    assert d["GEN_ATTRIBUTE_VALUE4"] == "SUP9"    # Supplier
    assert d["GEN_ATTRIBUTE_VALUE10"] == "IT1"    # Item Id
    assert d["GEN_ATTRIBUTE_VALUE11"] == "INV1"   # Invoice Number1


def test_load_id_qr_sheet():
    res = _mk(
        {"Material": ["X1"], "Material Description": ["d"], "Req Qty.": [2], "Delivery No": [555]},
        {"Material code": ["X1"], "Material Desc": ["d"], "Catergory": ["SET"], "HJ": [2], "SAP": [1]},
        {"Item Number": ["X1", "X1"], "Actual Qty": [2, 2], "Cbm": [0.1, 0.1]},
    )
    wb = load_workbook(io.BytesIO(res["vip_pick_bytes"]))
    ws = wb["LOAD ID QR"]
    assert ws["A1"].value == "LOAD ID"
    assert ws["D1"].value == "INM0VIP"


def test_load_id_dedup():
    args = (
        {"Material": ["X1"], "Material Description": ["d"], "Req Qty.": [2], "Delivery No": [777]},
        {"Material code": ["X1"], "Material Desc": ["d"], "Catergory": ["SET"], "HJ": [2], "SAP": [1]},
        {"Item Number": ["X1", "X1"], "Actual Qty": [2, 2], "Cbm": [0.1, 0.1]},
    )
    req, sku, inv = pd.DataFrame(args[0]), pd.DataFrame(args[1]), pd.DataFrame(args[2])
    res1 = pe.run_pipeline(req, sku, inv, pe.EngineConfig())
    existing = set(str(x) for x in res1["load_ids"])
    res2 = pe.run_pipeline(req, sku, inv, pe.EngineConfig(), existing_load_ids=existing)
    assert all("-" in str(x) for x in res2["load_ids"])


if __name__ == "__main__":
    test_set_multiplier_example()
    test_loose_equals_req()
    test_set_picks_whole_set()
    test_short_stock()
    test_not_in_sku_is_nodata()
    test_cannot_exceed_available()
    test_files_generate_and_format()
    test_india_detail_equals_pcs()
    test_gen_attributes_from_inventory()
    test_load_id_qr_sheet()
    test_load_id_dedup()
    print("All validation checks passed.")
